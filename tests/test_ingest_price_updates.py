"""
tests/test_ingest_price_updates.py
-----------------------------------
ingest_price_updates.py icin temel regresyon testi.

Test eder:
  1. Dry-run modunda Excel degismiyor (dosya boyutu ayni)
  2. --apply modunda price_tl_current kolonu ekleniyor
  3. --apply modunda fiyat dogru yaziliyor
  4. Backup dosyasi olusturuluyor

Calistirma:
    cd <repo_koku>
    python -m pytest tests/test_ingest_price_updates.py -v
    veya
    python tests/test_ingest_price_updates.py
"""

import json
import os
import shutil
import sys
import tempfile
import unittest

# Repo kokunu sys.path'e ekle
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from ingest_price_updates import main as ingest_price_main


REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_EXCEL = os.path.join(REPO_ROOT, "oyuncu_veritabani.xlsx")


class TestIngestPriceUpdates(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """Test icin gecici Excel kopyasi olustur."""
        cls.tmpdir = tempfile.mkdtemp(prefix="tff_test_")
        cls.excel_path = os.path.join(cls.tmpdir, "test_oyuncu.xlsx")
        shutil.copy2(SOURCE_EXCEL, cls.excel_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def _make_json(self, gameweek, prices):
        path = os.path.join(self.tmpdir, f"fiyat_gw{gameweek}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"gameweek": gameweek, "prices": prices}, f, ensure_ascii=False)
        return path

    def test_1_dry_run_does_not_modify_excel(self):
        """Dry-run modunda Excel dosyasi DEGISMEMELI."""
        json_path = self._make_json(1, [
            {"player_name": "Ebere Paul Onuachu", "team": "Trabzonspor", "price_tl": 9999999}
        ])
        size_before = os.path.getsize(self.excel_path)

        # Dry-run: --apply yok
        sys.argv = ["ingest_price_updates.py", self.excel_path, json_path]
        ingest_price_main()

        size_after = os.path.getsize(self.excel_path)
        self.assertEqual(size_before, size_after,
                         "Dry-run Excel'i degistirdi (boyut farki)")

    def test_2_apply_adds_price_column(self):
        """--apply modunda price_tl_current kolonu eklenmeli."""
        json_path = self._make_json(2, [
            {"player_name": "Ebere Paul Onuachu", "team": "Trabzonspor", "price_tl": 11000000}
        ])

        sys.argv = ["ingest_price_updates.py", self.excel_path, json_path, "--apply"]
        ingest_price_main()

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Players"]
        headers_after = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        wb.close()
        self.assertIn("price_tl_current", headers_after,
                      "price_tl_current kolonu eklenmedi")

    def test_3_apply_writes_correct_price(self):
        """--apply modunda fiyat dogru yazilmali."""
        test_price = 12345678
        json_path = self._make_json(3, [
            {"player_name": "Ebere Paul Onuachu", "team": "Trabzonspor", "price_tl": test_price}
        ])

        sys.argv = ["ingest_price_updates.py", self.excel_path, json_path, "--apply"]
        ingest_price_main()

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Players"]
        header_row = None
        for r in range(1, 10):
            if ws.cell(row=r, column=1).value == "player_id":
                header_row = r
                break
        headers = {ws.cell(row=header_row, column=c).value: c
                   for c in range(1, ws.max_column + 1)}
        price_col = headers["price_tl_current"]
        name_col = headers["name"]

        written_price = None
        r = header_row + 1
        while ws.cell(row=r, column=1).value:
            name = ws.cell(row=r, column=name_col).value
            if name and "Onuachu" in str(name):
                written_price = ws.cell(row=r, column=price_col).value
                break
            r += 1
        wb.close()

        self.assertIsNotNone(written_price, "Onuachu'nun fiyati bulunamadi")
        self.assertEqual(int(written_price), test_price,
                         f"Fiyat {test_price} yerine {written_price} yazildi")

    def test_4_backup_created_on_apply(self):
        """--apply modunda backups/ klasorunde yedek olusmali."""
        json_path = self._make_json(4, [
            {"player_name": "Ebere Paul Onuachu", "team": "Trabzonspor", "price_tl": 8000000}
        ])

        backups_dir = os.path.join(self.tmpdir, "backups")
        if os.path.exists(backups_dir):
            shutil.rmtree(backups_dir)

        sys.argv = ["ingest_price_updates.py", self.excel_path, json_path, "--apply"]
        ingest_price_main()

        self.assertTrue(os.path.exists(backups_dir),
                        "backups/ klasoru olusmadi")
        backup_files = [f for f in os.listdir(backups_dir) if f.endswith(".xlsx")]
        self.assertGreater(len(backup_files), 0,
                           "backups/ klasorunde .xlsx yedek yok")


if __name__ == "__main__":
    unittest.main(verbosity=2)
