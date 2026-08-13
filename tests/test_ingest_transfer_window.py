"""
tests/test_ingest_transfer_window.py
-------------------------------------
ingest_transfer_window.py icin temel regresyon testi.

Test eder:
  1. is_active kolonu ilk --apply'da ekleniyor
  2. "in" tipinde yeni oyuncu ekleniyor (player_id PLY444+ devam ediyor)
  3. "out" tipinde is_active=0 yapiliyor (satir SILINMIYOR)
  4. Birden fazla "in" transferinde player_id'ler benzersiz

Calistirma:
    python tests/test_ingest_transfer_window.py
"""

import json
import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from ingest_transfer_window import main as ingest_transfer_main


REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_EXCEL = os.path.join(REPO_ROOT, "oyuncu_veritabani.xlsx")


class TestIngestTransferWindow(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmpdir = tempfile.mkdtemp(prefix="tff_test_transfer_")
        cls.excel_path = os.path.join(cls.tmpdir, "test_oyuncu.xlsx")
        shutil.copy2(SOURCE_EXCEL, cls.excel_path)

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir, ignore_errors=True)

    def _make_json(self, transfers, transfer_date="2026-01-15"):
        path = os.path.join(self.tmpdir, "transfer.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"transfer_date": transfer_date,
                       "transfers": transfers}, f, ensure_ascii=False)
        return path

    def _get_headers(self, ws):
        for r in range(1, 10):
            if ws.cell(row=r, column=1).value == "player_id":
                header_row = r
                break
        return {ws.cell(row=header_row, column=c).value: c
                for c in range(1, ws.max_column + 1)}, header_row

    def test_1_is_active_column_added(self):
        """İlk --apply'da is_active kolonu eklenmeli, tüm oyuncular=1."""
        json_path = self._make_json([
            {"player_name": "Yeni Oyuncu X", "transfer_type": "in",
             "team": "Galatasaray", "position": "DEF - Defans",
             "new_price_tl": 5000000, "source_note": "test"}
        ])

        sys.argv = ["ingest_transfer_window.py", self.excel_path, json_path, "--apply"]
        ingest_transfer_main()

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Players"]
        headers, header_row = self._get_headers(ws)
        self.assertIn("is_active", headers, "is_active kolonu eklenmedi")

        # İlk 5 oyuncunun is_active=1 olmalı
        is_active_col = headers["is_active"]
        n_checked = 0
        r = header_row + 1
        while ws.cell(row=r, column=1).value and n_checked < 5:
            val = ws.cell(row=r, column=is_active_col).value
            self.assertEqual(int(val), 1,
                             f"Satir {r}: is_active=1 olmali, {val}")
            n_checked += 1
            r += 1
        wb.close()

    def test_2_new_player_added_with_correct_id(self):
        """Yeni oyuncu eklendiğinde player_id PLY444+ formatında olmalı."""
        # Mevcut max ID'yi bul
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Players"]
        headers, header_row = self._get_headers(ws)
        existing_ids = []
        r = header_row + 1
        while ws.cell(row=r, column=1).value:
            existing_ids.append(ws.cell(row=r, column=1).value)
            r += 1
        wb.close()

        import re
        max_n = 0
        for pid in existing_ids:
            m = re.match(r"^PLY(\d+)$", str(pid))
            if m:
                n = int(m.group(1))
                if n > max_n:
                    max_n = n

        json_path = self._make_json([
            {"player_name": "Test Transfer Oyuncusu", "transfer_type": "in",
             "team": "Fenerbahce", "position": "FWD - Forvet",
             "new_price_tl": 7000000, "source_note": "test"}
        ])

        sys.argv = ["ingest_transfer_window.py", self.excel_path, json_path, "--apply"]
        ingest_transfer_main()

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Players"]
        headers, header_row = self._get_headers(ws)
        name_col = headers["name"]
        id_col = headers["player_id"]

        found_id = None
        r = header_row + 1
        while ws.cell(row=r, column=1).value:
            name = ws.cell(row=r, column=name_col).value
            if name and "Test Transfer" in str(name):
                found_id = ws.cell(row=r, column=id_col).value
                break
            r += 1
        wb.close()

        self.assertIsNotNone(found_id, "Yeni oyuncu eklenmedi")
        m = re.match(r"^PLY(\d+)$", str(found_id))
        self.assertIsNotNone(m, f"player_id formati yanlis: {found_id}")
        self.assertGreater(int(m.group(1)), max_n,
                           f"Yeni ID {found_id} mevcut max PLY{max_n}'ten buyuk olmali")

    def test_3_out_does_not_delete_row(self):
        """'out' transferi satiri SILMEMELI, is_active=0 yapmali."""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Players"]
        headers, header_row = self._get_headers(ws)
        rows_before = 0
        r = header_row + 1
        while ws.cell(row=r, column=1).value:
            rows_before += 1
            r += 1
        wb.close()

        # Bilinen bir oyuncuyu pasifleştir (Mbaye Diagne - PLY443 olmali)
        json_path = self._make_json([
            {"player_name": "Mbaye Diagne", "transfer_type": "out",
             "team": "Amed", "source_note": "test"}
        ])

        sys.argv = ["ingest_transfer_window.py", self.excel_path, json_path, "--apply"]
        ingest_transfer_main()

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Players"]
        headers, header_row = self._get_headers(ws)
        rows_after = 0
        is_active_col = headers["is_active"]
        name_col = headers["name"]
        diagne_active = None
        r = header_row + 1
        while ws.cell(row=r, column=1).value:
            rows_after += 1
            name = ws.cell(row=r, column=name_col).value
            if name and "Diagne" in str(name):
                diagne_active = ws.cell(row=r, column=is_active_col).value
            r += 1
        wb.close()

        self.assertEqual(rows_after, rows_before,
                         "Satir sayisi degisti - 'out' satir silmemeli")
        self.assertEqual(int(diagne_active), 0,
                         "Diagne'nin is_active=0 olmali")

    def test_4_multiple_in_transfers_have_unique_ids(self):
        """Birden fazla 'in' transferinde player_id'ler benzersiz olmalı."""
        json_path = self._make_json([
            {"player_name": "Coklu Transfer 1", "transfer_type": "in",
             "team": "Besiktas", "position": "MID - Orta Saha",
             "new_price_tl": 4000000, "source_note": "test"},
            {"player_name": "Coklu Transfer 2", "transfer_type": "in",
             "team": "Trabzonspor", "position": "DEF - Defans",
             "new_price_tl": 3500000, "source_note": "test"},
        ])

        sys.argv = ["ingest_transfer_window.py", self.excel_path, json_path, "--apply"]
        ingest_transfer_main()

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Players"]
        headers, header_row = self._get_headers(ws)
        name_col = headers["name"]
        id_col = headers["player_id"]

        new_ids = []
        r = header_row + 1
        while ws.cell(row=r, column=1).value:
            name = ws.cell(row=r, column=name_col).value
            if name and "Coklu Transfer" in str(name):
                new_ids.append(ws.cell(row=r, column=id_col).value)
            r += 1
        wb.close()

        self.assertEqual(len(new_ids), 2,
                         f"2 yeni oyuncu eklenmeli, {len(new_ids)} bulundu")
        self.assertNotEqual(new_ids[0], new_ids[1],
                            f"player_id'ler benzersiz olmali: {new_ids}")


if __name__ == "__main__":
    unittest.main(verbosity=2)
