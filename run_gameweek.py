"""
run_gameweek.py
----------------
Haftalik calistirilacak ana optimizasyon ve kadro uretim script'i.
"""

import argparse
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from data_loader import load_players, load_gameweek_log
from xp_model import compute_xp
from optimizer import optimize_squad

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CAPTAIN_FILL = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
VICE_FILL    = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
BENCH_FILL   = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
THIN_BORDER  = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)

POS_ORDER = {"GK": 1, "DEF": 2, "MID": 3, "FWD": 4}
ROLE_ORDER = {"CAPTAIN": 1, "VICE_CAPTAIN": 2, "STARTER": 3, "BENCH": 4}


def format_tl(x: float) -> str:
    return f"{x:,.0f} TL".replace(",", ".")


def print_result(result, gameweek: int):
    print(f"\n{'='*65}")
    print(f"  HAFTA {gameweek} - EN IYI KADRO (Taktik Dizilis: {result.formation})")
    print(f"{'='*65}")
    print(f"Toplam Butce Kullanimi: {format_tl(result.total_budget_used)} / 100.000.000 TL")
    print(f"Ilk 11 Toplam xP: {result.total_starting_xp:.2f} Puan (+ Kaptan: {result.captain['xp']:.2f})")

    print("\n--- ILK 11 ---")
    for pos in ["GK", "DEF", "MID", "FWD"]:
        rows = result.starters[result.starters["position_code"] == pos]
        for _, r in rows.sort_values("xp", ascending=False).iterrows():
            tag = ""
            if r["player_id"] == result.captain["player_id"]:
                tag = "  <- ★ KAPTAN (2x)"
            elif r["player_id"] == result.vice_captain["player_id"]:
                tag = "  <- YEDEK KAPTAN"
            opp_str = f"(vs {r.get('fdr_opp', '')})" if r.get('fdr_opp') else ""
            print(f"  [{pos}] {r['name']:<28} {r['team']:<18} "
                  f"{format_tl(r['price_tl']):>14}  xP={r['xp']:.2f} {opp_str:<18}{tag}")

    print("\n--- YEDEKLER (oncelik sirasina gore) ---")
    if result.bench_gk is not None:
        print(f"  [GK]  {result.bench_gk['name']:<28} {result.bench_gk['team']:<18} {format_tl(result.bench_gk['price_tl']):>14}  xP={result.bench_gk['xp']:.2f}")
    for i, (_, r) in enumerate(result.bench_outfield.iterrows(), start=1):
        print(f"  [{i}]   {r['name']:<28} ({r['position_code']}) {r['team']:<14} {format_tl(r['price_tl']):>14}  xP={r['xp']:.2f}")


def export_result(result, path: str, gameweek: int):
    squad = result.squad.copy()
    squad["gameweek"] = gameweek
    squad["role"] = squad["is_starter"].map({True: "STARTER", False: "BENCH"})
    squad.loc[squad["player_id"] == result.captain["player_id"], "role"] = "CAPTAIN"
    squad.loc[squad["player_id"] == result.vice_captain["player_id"], "role"] = "VICE_CAPTAIN"

    out_cols = ["player_id", "name", "position_code", "team", "price_tl",
                "play_probability", "xp", "role"]
    if "fdr_opp" in squad.columns:
        out_cols.append("fdr_opp")

    squad_out = squad[out_cols].copy()

    squad_out["_pos_order"] = squad_out["position_code"].map(POS_ORDER)
    squad_out["_role_order"] = squad_out["role"].map(ROLE_ORDER)
    squad_out = squad_out.sort_values(
        ["_pos_order", "_role_order", "xp"], ascending=[True, True, False]
    ).drop(columns=["_pos_order", "_role_order"]).reset_index(drop=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"GW{gameweek}_Kadro"

    ws.append([f"TFF FANTEZİ LİGİ - HAFTA {gameweek} KADRO ÖNERİSİ (Diziliş: {result.formation})"])
    ws.append([f"Toplam Maliyet: {format_tl(result.total_budget_used)} | İlk 11 xP: {result.total_starting_xp:.2f}"])
    ws.append([])

    headers = ["Oyuncu ID", "Ad", "Pozisyon", "Takım", "Fiyat", "Oynama %", "xP", "Rol"]
    if "fdr_opp" in out_cols:
        headers.append("Haftanın Rakibi")
    ws.append(headers)

    for _, r in squad_out.iterrows():
        row_vals = [
            r["player_id"], r["name"], r["position_code"], r["team"],
            r["price_tl"], r["play_probability"], r["xp"], r["role"]
        ]
        if "fdr_opp" in out_cols:
            row_vals.append(r.get("fdr_opp", ""))
        ws.append(row_vals)

    for cell in ws[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        role_val = row[7].value
        fill_to_apply = None
        if role_val == "CAPTAIN":
            fill_to_apply = CAPTAIN_FILL
        elif role_val == "VICE_CAPTAIN":
            fill_to_apply = VICE_FILL
        elif role_val == "BENCH":
            fill_to_apply = BENCH_FILL

        for cell in row:
            cell.border = THIN_BORDER
            if fill_to_apply:
                cell.fill = fill_to_apply

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            if len(val) > max_len and cell.row > 2:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(path)
    print(f"\nSonuc yazildi: {path}")


def main():
    parser = argparse.ArgumentParser(description="Haftalik TFF Fantezi Lig kadro optimizasyonu")
    parser.add_argument("excel_path", help="Excel veritabani dosya yolu")
    parser.add_argument("--gameweek", type=int, default=2, help="Hafta numarasi")
    parser.add_argument("--out", default=None, help="Cikti Excel dosyasi yolu")
    args = parser.parse_args()

    out_path = args.out or f"gw{args.gameweek}_kadro_onerisi.xlsx"

    fixtures_json = f"nostradamus_predict_gw{args.gameweek}.json"
    if not os.path.exists(fixtures_json):
        fixtures_json = f"nostradamus_fixtures_gw{args.gameweek}.json"

    players = load_players(args.excel_path)
    log = load_gameweek_log(args.excel_path)

    df_xp = compute_xp(players, log, fixtures_path=fixtures_json)
    result = optimize_squad(df_xp)

    print_result(result, args.gameweek)
    export_result(result, out_path, args.gameweek)


if __name__ == "__main__":
    main()
