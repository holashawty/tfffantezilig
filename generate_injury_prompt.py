"""
generate_injury_prompt.py
-------------------------
Excel veritabanındaki (oyuncu_veritabani.xlsx) 462 oyuncunun tamamını
güncel kulüplerine göre gruplayarak, LLM halüsinasyonlarını (eski takım,
eski sezon sakatlıkları) ve gözden kaçan oyuncuları %100 engelleyen
özel ve dinamik bir araştırma promptu üretir.

Kullanım:
    python generate_injury_prompt.py [--gameweek 2] [--excel oyuncu_veritabani.xlsx]
"""

import argparse
import os
from datetime import datetime
import openpyxl


def generate_prompt(excel_path="oyuncu_veritabani.xlsx", gameweek=2, out_path="web_arastirma_prompti.md"):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Players"]

    header_row = None
    for r in range(1, 10):
        if ws.cell(row=r, column=1).value == "player_id":
            header_row = r
            break

    if not header_row:
        raise RuntimeError("player_id header satırı bulunamadı.")

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            headers[v] = c

    # Takımlara göre oyuncuları topla
    teams_dict = {}
    total_players = 0
    r = header_row + 1
    while ws.cell(row=r, column=headers["player_id"]).value:
        name = ws.cell(row=r, column=headers["name"]).value
        team = ws.cell(row=r, column=headers["team"]).value
        pos = ws.cell(row=r, column=headers["position"]).value
        is_active = ws.cell(row=r, column=headers["is_active"]).value if "is_active" in headers else 1

        if is_active in (1, "1", None) and name and team:
            if team not in teams_dict:
                teams_dict[team] = []
            teams_dict[team].append(name)
            total_players += 1
        r += 1

    today_str = datetime.now().strftime("%d.%m.%Y")
    today_iso = datetime.now().strftime("%Y-%m-%d")

    prompt_content = f"""# Haftalık Sakatlık & Ceza Araştırma Promptu (Hafta {gameweek} - {today_str})

Bu prompt, veritabanındaki **{total_players} oyuncunun tamamını** ve kulüplerini içerir.
Web araması yapabilen bir AI'ya (Gemini, ChatGPT, Claude) aşağıdaki kutunun tamamını kopyalayıp yapıştırın.

---

## 📋 KOPYALA-YAPISTIR PROMPT:

```text
GÖREV: Türkiye Trendyol Süper Lig 2026-2027 Sezonu {gameweek}. Hafta maçları öncesi sakatlık, ceza ve kadro dışı durumlarını araştır.
GÜNCEL TARİH: {today_str}

KRİTİK ZAMAN VE HALÜSİNASYON KISITLARI (KESİN KURAL):
1. SADECE son 7-10 gün içindeki ({datetime.now().strftime("%B %Y")}) GÜNCEL haberleri ve resmi açıklamaları baz al. 2024, 2025 veya geçmiş aylardaki eski sakatlık haberlerini KESİNLİKLE DAHİL ETME.
2. Aşağıda 18 Süper Lig kulübü ve bu kulüplerin 2026-2027 sezonu GÜNCEL oyuncu kadroları verilmiştir. Oyuncuların takımları BU LİSTEYLE SABİTTİR. (Örn: Anthony Musaba Fenerbahçe'dedir). Oyuncuların eski takımlarına dair haberlerle karıştırma.
3. SADECE oynamayacak, sakat, cezalı veya şüpheli oyuncuları listele. Sağlam/oynamaya hazır oyuncuları ekleme.
4. Uydurma bilgi verme; her sakatlık için kesin haber kaynağı ve tarih belirt (Örn: PFDK 1 maç ceza / Kulüp Sağlık Bülteni {today_str}).

AŞAĞIDAKİ 18 KULÜBÜ VE KADROLARINI TEK TEK TARA:
"""

    for team_name in sorted(teams_dict.keys()):
        players = teams_dict[team_name]
        players_joined = ", ".join(players)
        prompt_content += f"\n[{team_name.upper()} ({len(players)} Oyuncu)]:\n{players_joined}\n"

    prompt_content += f"""
ÇIKTI FORMATI:
SADECE aşağıdaki JSON formatında, JSON DIŞINDA HİÇBİR AÇIKLAMA EKLEMEDEN cevap ver:

{{
  "gameweek": {gameweek},
  "research_date": "{today_iso}",
  "updates": [
    {{
      "player_name": "Oyuncunun Listedeki Tam Adı",
      "team": "Takım Adı",
      "status": "injured | suspended | doubtful | rotation_risk",
      "play_probability": 0.0,
      "source_note": "Haber kaynağı ve tarih (Örn: Kulüp Basın Bülteni {today_str})"
    }}
  ]
}}

play_probability Rehberi:
- Kesin oynamayacak (sakat / cezalı / kadro dışı): 0.0
- Ciddi şüpheli / maç günü testi: 0.25
- Hafif sakatlık atlatan / sonradan girme riski: 0.50
- Oynaması beklenen ama riski olan: 0.75
```

---

## 💾 Dosyayı Kaydetme

Çıkan JSON yanıtını `web_research_gw{gameweek}.json` olarak kaydedin ve `rehber.bat` menüsünden `[2] -> [x] Hızlı Mod` ile sisteme yükleyin.
"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(prompt_content)

    print(f"[+] Başarılı! {len(teams_dict)} takım ve {total_players} oyuncu içeren dinamik prompt oluşturuldu: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Dinamik Sakatlık Promptu Üretici")
    parser.add_argument("--gameweek", type=int, default=2, help="Hafta numarası")
    parser.add_argument("--excel", default="oyuncu_veritabani.xlsx", help="Excel veritabanı")
    parser.add_argument("--out", default="web_arastirma_prompti.md", help="Çıktı markdown dosyası")
    args = parser.parse_args()

    generate_prompt(args.excel, args.gameweek, args.out)


if __name__ == "__main__":
    main()
