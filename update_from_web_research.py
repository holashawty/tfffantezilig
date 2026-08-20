"""
update_from_web_research.py
----------------------------
web_arastirma_prompti.md ile bir web AI'dan alinan JSON ciktisini
Excel'deki Players sheet'inin play_probability kolonuna isler.

Neden dogrudan otomatik degil de "goster + onayla" akisi:
Web AI'nin arastirmasi yanlis/eksik/uydurma olabilir. Bu script hicbir
guncellemeyi SESSIZCE uygulamaz — her satiri ekrana yazar, eslesmeyen
veya belirsiz isimleri ayri listeler. Kor guven yok.

Kullanim:
    python update_from_web_research.py <excel_yolu> <web_research_gwN.json> [--apply]

    --apply verilmezse SADECE rapor gosterir, dosyaya yazmaz (dry-run).
    --apply verilirse Excel'i gunceller.
"""

import argparse
import json
import sys
from difflib import SequenceMatcher

import openpyxl

from validator import validate_injury_updates
from backup_utils import backup_excel, safe_save_excel

NAME_MATCH_THRESHOLD = 0.82  # bunun altinda eslesme "belirsiz" sayilir


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def load_players_rows(ws):
    """Header satirini bulur, {row_index: {player_id, name, team}} doner."""
    header_row = None
    for r in range(1, 10):
        if ws.cell(row=r, column=1).value == "player_id":
            header_row = r
            break
    if header_row is None:
        raise RuntimeError("player_id header satiri bulunamadi.")

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            headers[v] = c

    required = ["player_id", "name", "team", "play_probability"]
    for col in required:
        if col not in headers:
            raise RuntimeError(f"Beklenen kolon yok: {col}")

    rows = {}
    r = header_row + 1
    while ws.cell(row=r, column=headers["player_id"]).value:
        rows[r] = {
            "player_id": ws.cell(row=r, column=headers["player_id"]).value,
            "name": ws.cell(row=r, column=headers["name"]).value,
            "team": ws.cell(row=r, column=headers["team"]).value,
        }
        r += 1
    return headers, rows


def match_update(update, rows):
    """En iyi isim eslesmesini bulur. (row_idx, score) ya da (None, 0) doner.
    Gelistirilmis akilli eslesme: Tam eslesme, fuzzy eslesme ve alt-kelime (Mert Gunok -> Fehmi Mert Gunok) destegi."""
    best_row, best_score = None, 0.0
    
    u_name = update["player_name"].lower().strip()
    u_words = set(u_name.split())
    u_team = (update.get("team") or "").lower().strip()

    for row_idx, info in rows.items():
        db_name = info["name"].lower().strip()
        db_words = set(db_name.split())
        db_team = (info.get("team") or "").lower().strip()
        
        team_matches = False
        if u_team and db_team:
            if u_team in db_team or db_team in u_team:
                team_matches = True
                
        # 1. Standart benzerlik
        score = _similarity(u_name, db_name)
        if team_matches:
            score += 0.05
            
        # 2. Alt-kelime kontrolu (Orn: "Mert Gunok" -> "Fehmi Mert Gunok", "David Costa" -> "David Jose Soares... Costa")
        if u_words and u_words.issubset(db_words):
            subset_score = 0.90 + (0.10 if team_matches else 0.0)
            if subset_score > score:
                score = subset_score
                
        # Soyad + ilk isim uyumu
        if len(u_words) >= 2 and team_matches:
            u_first, u_last = list(u_words)[0], list(u_words)[-1]
            if u_last in db_words:
                score = max(score, 0.85)

        if score > best_score:
            best_row, best_score = row_idx, score
            
    return best_row, best_score


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel_path")
    ap.add_argument("json_path")
    ap.add_argument("--apply", action="store_true",
                     help="Verilmezse sadece rapor gosterir, dosyayi degistirmez.")
    args = ap.parse_args()

    with open(args.json_path, "r", encoding="utf-8") as f:
        research = json.load(f)

    wb = openpyxl.load_workbook(args.excel_path)
    ws = wb["Players"]
    headers, rows = load_players_rows(ws)

    # ADIM 1: veri kalitesi kontrolu (once matching degil, once gecerlilik)
    quality = validate_injury_updates(research.get("updates", []), total_player_count=len(rows))
    quality.print_report(f"Hafta {research.get('gameweek')} sakatlik/ceza")

    matched, unmatched = [], []

    for upd in quality.valid_records:
        row_idx, score = match_update(upd, rows)
        if row_idx is None or score < NAME_MATCH_THRESHOLD:
            unmatched.append(upd)
            continue
        info = rows[row_idx]
        matched.append({
            "row": row_idx,
            "player_id": info["player_id"],
            "matched_name": info["name"],
            "input_name": upd["player_name"],
            "similarity": round(score, 3),
            "status": upd.get("status"),
            "old_prob": ws.cell(row=row_idx, column=headers["play_probability"]).value,
            "new_prob": upd["play_probability"],
            "source_note": upd.get("source_note", ""),
        })

    print(f"\n=== Hafta {research.get('gameweek')} eslesme raporu ===")
    print(f"Gecerli kayittan eslesen: {len(matched)}  |  Isim eslesmeyen: {len(unmatched)}\n")

    print("--- ESLESENLER (uygulanacak) ---")
    for m in matched:
        print(f"  {m['input_name']:<28} -> {m['matched_name']:<28} "
              f"[{m['status']}] {m['old_prob']} -> {m['new_prob']}  "
              f"(benzerlik={m['similarity']})  {m['source_note']}")

    if unmatched:
        print("\n--- ESLESMEYEN (MANUEL KONTROL ET, otomatik uygulanmadi) ---")
        for u in unmatched:
            print(f"  {u['player_name']} ({u.get('team')}) - {u.get('source_note','')}")

    if not args.apply:
        print("\n[DRY-RUN] Hicbir sey yazilmadi. Onayliyorsan --apply ile tekrar calistir.")
        return

    # --apply oncesi Excel yedegi al (docs/07 13 Agustos 2026 karari)
    backup_excel(args.excel_path)

    for m in matched:
        ws.cell(row=m["row"], column=headers["play_probability"]).value = m["new_prob"]

    safe_save_excel(wb, args.excel_path)
    print(f"\n[UYGULANDI] {len(matched)} satir guncellendi -> {args.excel_path}")


if __name__ == "__main__":
    main()
