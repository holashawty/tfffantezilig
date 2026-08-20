"""
ingest_price_updates.py
-------------------------
Transfer piyasasindaki GUNCEL fiyatlari (price_tl_current) Players
sheet'ine isler. Fiyatlar dinamik oldugu icin (TFF'nin kendi ic
ekonomisi) bu, play_probability ile ayni mantikta CANLI bir alan —
GameweekLog'a degil, dogrudan Players'a yazilir (bkz. data_loader.py
load_players() aciklamasi).

Kolon otomatik olusturulur: Players sheet'inde price_tl_current yoksa
bu script once onu ekler (baslangicta price_tl_gw1 ile ayni deger),
sonra sadece JSON'da gelen guncellemeleri uygular.

JSON semasi (fiyat_guncelleme_prompti.md bunu uretir):
{
  "gameweek": 2,
  "prices": [
    {"player_name": "...", "team": "...", "price_tl": 5500000}
  ]
}

Kullanim:
    python ingest_price_updates.py <excel_yolu> <fiyat_gwN.json> [--apply]
"""

import argparse
import json
from difflib import SequenceMatcher

import openpyxl

from validator import validate_price_updates
from backup_utils import backup_excel, safe_save_excel

NAME_MATCH_THRESHOLD = 0.82


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _get_or_create_headers(ws):
    header_row = None
    for r in range(1, 10):
        if ws.cell(row=r, column=1).value == "player_id":
            header_row = r
            break
    if header_row is None:
        raise RuntimeError("player_id header satiri bulunamadi.")

    headers = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            headers[v] = c

    if "price_tl_current" not in headers:
        new_col = max_col + 1
        ws.cell(row=header_row, column=new_col).value = "price_tl_current"
        headers["price_tl_current"] = new_col
        # baslangic degeri: price_tl_gw1 ile ayni (henuz degisim yok varsayimi)
        base_col = headers.get("price_tl_gw1")
        r = header_row + 1
        while ws.cell(row=r, column=headers["player_id"]).value:
            if base_col:
                ws.cell(row=r, column=new_col).value = ws.cell(row=r, column=base_col).value
            r += 1
        print("[BILGI] Players sheet'ine 'price_tl_current' kolonu eklendi "
              "(baslangicta price_tl_gw1 ile ayni).")

    return header_row, headers


def load_rows(ws, header_row, headers):
    rows = {}
    r = header_row + 1
    while ws.cell(row=r, column=headers["player_id"]).value:
        rows[r] = {
            "player_id": ws.cell(row=r, column=headers["player_id"]).value,
            "name": ws.cell(row=r, column=headers["name"]).value,
            "team": ws.cell(row=r, column=headers["team"]).value,
        }
        r += 1
    return rows


def match_price(update, rows):
    best_row, best_score = None, 0.0
    u_name = update["player_name"].lower().strip()
    u_words = set(u_name.split())
    u_team = (update.get("team") or "").lower().strip()

    for row_idx, info in rows.items():
        db_name = info["name"].lower().strip()
        db_words = set(db_name.split())
        db_team = (info.get("team") or "").lower().strip()
        
        team_matches = False
        if u_team and db_team:
            if u_team in db_team or db_team in u_team:
                team_matches = True

        score = _similarity(u_name, db_name)
        if team_matches:
            score += 0.05

        if u_words and u_words.issubset(db_words):
            subset_score = 0.90 + (0.10 if team_matches else 0.0)
            if subset_score > score:
                score = subset_score

        if score > best_score:
            best_row, best_score = row_idx, score
            
    return best_row, best_score


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel_path")
    ap.add_argument("json_path")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    with open(args.json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(args.excel_path)
    ws = wb["Players"]
    header_row, headers = _get_or_create_headers(ws)
    rows = load_rows(ws, header_row, headers)

    quality = validate_price_updates(data.get("prices", []))
    quality.print_report(f"Hafta {data.get('gameweek')} fiyat guncelleme")

    matched, unmatched = [], []
    for upd in quality.valid_records:
        row_idx, score = match_price(upd, rows)
        if row_idx is None or score < NAME_MATCH_THRESHOLD:
            unmatched.append(upd)
            continue
        info = rows[row_idx]
        old_price = ws.cell(row=row_idx, column=headers["price_tl_current"]).value
        matched.append({
            "row": row_idx, "matched_name": info["name"],
            "input_name": upd["player_name"], "similarity": round(score, 3),
            "old_price": old_price, "new_price": upd["price_tl"],
        })

    print(f"\n=== Hafta {data.get('gameweek')} eslesme raporu ===")
    print(f"Gecerli kayittan eslesen: {len(matched)}  |  Isim eslesmeyen: {len(unmatched)}\n")
    for m in matched:
        delta = ""
        if m["old_price"] is not None:
            diff = m["new_price"] - m["old_price"]
            if diff != 0:
                delta = f"  ({'+' if diff > 0 else ''}{diff:,.0f} TL)".replace(",", ".")
        print(f"  {m['input_name']:<28} -> {m['matched_name']:<28} "
              f"{m['old_price']} -> {m['new_price']}{delta}")
    if unmatched:
        print("\n--- ESLESMEYEN (manuel kontrol) ---")
        for u in unmatched:
            print(f"  {u['player_name']} ({u.get('team')})")

    if not args.apply:
        print("\n[DRY-RUN] Hicbir sey yazilmadi (kolon eklenmesi dahil). "
              "Onayliyorsan --apply ile tekrar calistir.")
        return

    # --apply oncesi Excel yedegi al (docs/07 13 Agustos 2026 karari)
    backup_excel(args.excel_path)

    for m in matched:
        ws.cell(row=m["row"], column=headers["price_tl_current"]).value = m["new_price"]

    safe_save_excel(wb, args.excel_path)
    print(f"\n[UYGULANDI] {len(matched)} fiyat guncellendi -> {args.excel_path}")


if __name__ == "__main__":
    main()
