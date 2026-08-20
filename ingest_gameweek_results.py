
def _norm(text: str) -> str:
    if not text:
        return ""
    # Unicode karakterleri ve aksanlari normalize et (Orn: Savić -> savic, Žužek -> zuzek, Černý -> cerny)
    t = text.lower().replace("ı", "i").replace("ğ", "g").replace("ü", "u").replace("ş", "s").replace("ö", "o").replace("ç", "c")
    nfd = unicodedata.normalize('NFD', t)
    return "".join(c for c in nfd if unicodedata.category(c) != 'Mn').strip()

ALIASES = {
    "talisca": "anderson souza conceicao",
    "show": "manuel luis da silva cafumana",
    "maestro": "antonio simao muanza",
    "shomu": "eldor shomurodov",
    "dia saba": "diaa sabi'a",
    "outtara": "kassoum ouattara",
    "charles": "charles-andre raux-yao",
    "pina": "wagner fabricio cardoso de pina",
    "savic": "stefan savic",
    "saviolo": "noah jose damesquita e melo saviolo",
    "varesanovic": "dal varesanovic",
    "goncalves": "diogo antonio cupido goncalves",
    "mebude": "adedire emmanuel awokoya-mebude",
    "jevtovic": "marko jevtovic",
    "stefan": "florin bogdan stefan",
    "kozlowski": "kacper szymon kozlowski",
    "zuzek": "zan zuzek",
    "guendouzi": "matteo elias kenzo guendouzi olie",
    "ake": "nathan benjamin ake",
    "operi": "christopher tea domoraud operi",
    "kaluzinski": "jakub kaluzinski",
    "haidara": "massadio haidara",
    "rodriguez": "martin vladimir rodriguez torrejon",
    "mujakic": "nihad mujakic",
    "gerxhaliu": "amar gerxhaliu",
    "orbanic": "matija orbanic",
    "cerny": "vaclav cerny",
    "fredy": "alfredo kulembe gomes ribeiro",
    "jacobs": "ismail joshua jakobs",
    "oh": "hyeon-gyu oh",
}

import unicodedata
"""
ingest_gameweek_results.py
---------------------------
Bir hafta oynandiktan sonra GERCEK performans verisini (dakika, gol,
asist, kart, ve varsa TFF'nin kendi hesapladigi fantasy_points) JSON'dan
okuyup GameweekLog'a EKLER (var olan satirlari degil, yeni satirlar).

IKI VERI KAYNAGI ONERISI (ikisi de ayni JSON semasina cevrilir):
  A) EN GUVENILIR: TFF Fantezi Lig uygulamasindaki "Puanim" ekranini
     ekran goruntusu/PDF olarak belge-erisimi olan bir AI'ya (Gemini)
     ver, "bu ekrandaki oyuncu puanlarini asagidaki JSON semasina
     donustur" de. fantasy_points dogrudan TFF'nin kendi hesabidir,
     bizim tahminimize gerek kalmaz.
  B) TAMAMLAYICI: Web arastirmasiyla (goal.com, sofascore vb.) mac
     istatistikleri (dakika/gol/asist/kart) toplanip PRIOR_CONFIG
     kalibrasyonu icin kullanilir — ama bu kaynak eksik/hatali
     olabilir, o yuzden fantasy_points alaniyla CELISIRSE (A) kaynagi
     esas alinir.

JSON semasi (match_sonuclari_prompti.md'deki prompt bunu uretir):
{
  "gameweek": 1,
  "results": [
    {
      "player_name": "...", "team": "...",
      "minutes": 90, "goals": 1, "assists": 0,
      "yellow_cards": 0, "red_cards": 0,
      "fantasy_points": 12.0,     // TFF'nin kendi puanindan biliniyorsa
      "price_tl": 4500000         // o haftaki fiyat biliniyorsa (opsiyonel)
    }
  ]
}

Kullanim:
    python ingest_gameweek_results.py <excel_yolu> <sonuclar_gwN.json> [--apply]
"""

import argparse
import json
from difflib import SequenceMatcher

import openpyxl

from data_loader import load_players
from validator import validate_match_results
from backup_utils import backup_excel, safe_save_excel

NAME_MATCH_THRESHOLD = 0.82


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio()


def _find_gameweeklog_header(ws):
    for r in range(1, 10):
        if ws.cell(row=r, column=1).value == "player_id":
            return r
    raise RuntimeError("GameweekLog header satiri bulunamadi.")


def _last_data_row(ws, header_row):
    r = header_row + 1
    while ws.cell(row=r, column=1).value not in (None, ""):
        r += 1
    return r  # ilk bos satir


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel_path")
    ap.add_argument("json_path")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    with open(args.json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    gameweek = data["gameweek"]
    results = data.get("results", [])

    players = load_players(args.excel_path)
    name_lookup = players[["player_id", "name", "team"]].to_dict("records")

    quality = validate_match_results(results)
    quality.print_report(f"Hafta {gameweek} mac sonuclari")

    matched, unmatched = [], []
    matched_ids = set()

    for res in quality.valid_records:
        best_id, best_score, best_name = None, 0.0, None
        raw_u_name = res["player_name"]
        norm_u = _norm(raw_u_name)
        # Alias kontrolu
        if norm_u in ALIASES:
            norm_u = ALIASES[norm_u]

        u_words = set(norm_u.split())
        norm_team = _norm(res.get("team") or "")

        for p in name_lookup:
            norm_db = _norm(p["name"])
            db_words = set(norm_db.split())
            db_team = _norm(p.get("team") or "")
            
            team_matches = False
            if norm_team and db_team:
                if norm_team in db_team or db_team in norm_team:
                    team_matches = True

            score = _similarity(norm_u, norm_db)
            if team_matches:
                score += 0.10

            # Alt kelime kontrolu (Orn: Osimhen in Victor James Osimhen)
            if u_words and any(w in norm_db for w in u_words if len(w) >= 3):
                subset_score = 0.88 + (0.12 if team_matches else 0.0)
                if subset_score > score:
                    score = subset_score

            if score > best_score:
                best_id, best_score, best_name = p["player_id"], score, p["name"]

        if best_id is None or best_score < NAME_MATCH_THRESHOLD:
            unmatched.append(res)
            continue

        matched_ids.add(best_id)
        matched.append({
            "player_id": best_id,
            "matched_name": best_name,
            "input_name": res["player_name"],
            "similarity": round(best_score, 3),
            "minutes": res.get("minutes", 0),
            "goals": res.get("goals", 0),
            "assists": res.get("assists", 0),
            "yellow_cards": res.get("yellow_cards", 0),
            "red_cards": res.get("red_cards", 0),
            "fantasy_points": res.get("fantasy_points", 0),
            "price_tl": res.get("price_tl"),
        })

    # Oynamayan diger futbolculari da 0 dakika / 0 puan olarak otomatik ekle (462 oyuncu butunlugu icin)
    for p in name_lookup:
        if p["player_id"] not in matched_ids:
            matched.append({
                "player_id": p["player_id"],
                "matched_name": p["name"],
                "input_name": "(Oynamadi / Kadro Disi)",
                "similarity": 1.0,
                "minutes": 0,
                "goals": 0,
                "assists": 0,
                "yellow_cards": 0,
                "red_cards": 0,
                "fantasy_points": 0,
                "price_tl": None,
            })

    print(f"\n=== Hafta {gameweek} eslesme raporu ===")
    print(f"Gecerli kayittan eslesen: {len(matched)}  |  Isim eslesmeyen: {len(unmatched)}\n")
    for m in matched:
        print(f"  {m['input_name']:<28} -> {m['matched_name']:<28} "
              f"dk={m['minutes']:>3} gol={m['goals']} asist={m['assists']} "
              f"puan={m['fantasy_points']}  (benzerlik={m['similarity']})")
    if unmatched:
        print("\n--- ESLESMEYEN (manuel kontrol) ---")
        for u in unmatched:
            print(f"  {u['player_name']} ({u.get('team')})")

    if not args.apply:
        print("\n[DRY-RUN] GameweekLog'a yazilmadi. Onayliyorsan --apply ile tekrar calistir.")
        return

    # --apply oncesi Excel yedegi al (docs/07 13 Agustos 2026 karari)
    backup_excel(args.excel_path)

    wb = openpyxl.load_workbook(args.excel_path)
    ws = wb["GameweekLog"]
    header_row = _find_gameweeklog_header(ws)
    r = _last_data_row(ws, header_row)

    for m in matched:
        ws.cell(row=r, column=1).value = m["player_id"]
        ws.cell(row=r, column=2).value = gameweek
        ws.cell(row=r, column=3).value = m["minutes"]
        ws.cell(row=r, column=4).value = m["goals"]
        ws.cell(row=r, column=5).value = m["assists"]
        ws.cell(row=r, column=6).value = m["yellow_cards"]
        ws.cell(row=r, column=7).value = m["red_cards"]
        ws.cell(row=r, column=8).value = m["fantasy_points"]
        if m["price_tl"] is not None:
            ws.cell(row=r, column=9).value = m["price_tl"]
        r += 1

    # SeasonStats sayfasini da otomatik guncelle
    try:
        from season_stats import compute_comprehensive_season_stats
        if "SeasonStats" not in wb.sheetnames:
            ws_stats = wb.create_sheet("SeasonStats")
        else:
            ws_stats = wb["SeasonStats"]
        
        ws_stats.delete_rows(1, ws_stats.max_row + 10)
        ws_stats.append(["TFF FANTEZİ LİG - SEZONLUK OYUNCU İSTATİSTİKLERİ VE LİDERLİK TABLOSU", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])
        ws_stats.append(["Tüm futbolcuların toplam maç, süre, puan, gol, asist, kart, form ve 90dk başına verimlilik metrikleri", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])
        ws_stats.append([None] * 18)
        stats_headers = [
            "player_id", "name", "team", "position_code", "price_tl",
            "total_games", "starts", "total_minutes", "avg_minutes",
            "total_points", "avg_points", "points_per_90", "total_goals",
            "total_assists", "total_yellow", "total_red", "total_saves", "form_score"
        ]
        ws_stats.append(stats_headers)
        
        # Temp save to compute with full log
        safe_save_excel(wb, args.excel_path)
        
        fresh_players = load_players(args.excel_path)
        fresh_log = load_gameweek_log(args.excel_path)
        season_df = compute_comprehensive_season_stats(fresh_players, fresh_log)
        season_df = season_df.sort_values(["total_points", "form_score", "price_tl"], ascending=[False, False, False])
        
        for _, row_s in season_df.iterrows():
            ws_stats.append([
                row_s["player_id"], row_s["name"], row_s["team"], row_s["position_code"], int(row_s["price_tl"]),
                int(row_s["total_games"]), int(row_s["starts"]), int(row_s["total_minutes"]), float(row_s["avg_minutes"]),
                int(row_s["total_points"]), float(row_s["avg_points"]), float(row_s["points_per_90"]), int(row_s["total_goals"]),
                int(row_s["total_assists"]), int(row_s["total_yellow"]), int(row_s["total_red"]), int(row_s["total_saves"]), float(row_s["form_score"])
            ])
        print("[BILGI] SeasonStats liderlik tablosu otomatik guncellendi.")
    except Exception as e:
        print(f"[UYARI] SeasonStats guncellenirken hata olustu: {e}")

    safe_save_excel(wb, args.excel_path)
    print(f"\n[UYGULANDI] {len(matched)} satir GameweekLog'a eklendi -> {args.excel_path}")


if __name__ == "__main__":
    main()
