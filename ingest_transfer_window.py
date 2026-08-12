"""
ingest_transfer_window.py
-------------------------
Sezon ortası transfer penceresi güncellemelerini Players sheet'ine işler.

İLK ÇALIŞTIRMADA: Players sheet'ine `is_active` kolonu otomatik eklenir
(tüm mevcut oyuncular için is_active=1). Bu kolon-ekleme deseni
`ingest_price_updates.py`'nin `price_tl_current` ekleme mantığıyla birebir
aynıdır — bkz. `_get_or_create_headers`.

İŞLEM TİPLERİ (bkz. validator.py `validate_transfer_window`):
  - "in":  Yeni oyuncu Süper Lig'e geldi → yeni player_id (PLY444, PLY445...)
           atanır, yeni satır eklenir. last_season_* alanları BOŞ (uydurma).
           Fiyat new_price_tl ile (verilmezse varsayılan 2.000.000 TL).
  - "out": Oyuncu Süper Lig'den ayrıldı → is_active=0. SATIR SİLİNMEZ
           (GameweekLog geçmiş referansları kırılır, bkz. docs/06).
  - "move": Oyuncu Süper Lig içinde takım değiştirdi → team alanı güncellenir,
           is_active değişmez.

GÜVENLİK DESENİ (docs/05'teki standart akış):
  1. JSON oku
  2. validator.py ile doğrula (zorunlu alan / değer aralık kontrolü)
  3. İsim eşleştir (fuzzy match, `difflib.SequenceMatcher`)
     — "in" tipinde: mevcut oyuncu VARSAYILAN olarak eşleşme beklenmez,
       ama önceki pasif kayıtla çakışma varsa uyarı ver
     — "out" ve "move" tiplerinde: en iyi eşleşmeyi bul
  4. Rapor bas (dry-run) — HİÇBİR ŞEY YAZILMAZ
  5. --apply verilirse gerçek yazma

JSON ŞEMASI (transfer_prompti.md ile üretilir):
{
  "transfer_date": "2026-01-15",
  "transfers": [
    {
      "player_name": "Ahmet Yılmaz",
      "transfer_type": "in",        # in | out | move
      "team": "Galatasaray",         # in/move için zorunlu
      "position": "DEF - Defans",    # in için zorunlu
      "new_price_tl": 5500000,       # in için opsiyonel
      "source_note": "transfermarkt.com"
    }
  ]
}

Kullanım:
    python ingest_transfer_window.py <excel_yolu> <transfer_window.json> [--apply]
"""

import argparse
import json
import re
from difflib import SequenceMatcher

import openpyxl

from validator import validate_transfer_window


NAME_MATCH_THRESHOLD = 0.82
DEFAULT_NEW_PLAYER_PRICE = 2_000_000  # yeni transfer için fiyat verilmezse


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, str(a).lower().strip(), str(b).lower().strip()).ratio()


def _next_player_id(existing_ids):
    """PLY443, PLY444, PLY445... formatında bir sonraki ID üretir.
    Mevcut en yüksek PLY<num> sayısını bulur, bir artırır."""
    max_n = 0
    for pid in existing_ids:
        m = re.match(r"^PLY(\d+)$", str(pid))
        if m:
            n = int(m.group(1))
            if n > max_n:
                max_n = n
    return f"PLY{max_n + 1}"


def _get_or_create_headers(ws):
    """`ingest_price_updates.py`'deki ile aynı desen.
    Header satırını bulur; `is_active` kolonu yoksa otomatik ekler
    (tüm mevcut oyuncular için is_active=1) ve headers dict'ini döndürür."""
    header_row = None
    for r in range(1, 10):
        if ws.cell(row=r, column=1).value == "player_id":
            header_row = r
            break
    if header_row is None:
        raise RuntimeError("player_id header satiri bulunamadi.")

    headers = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            headers[v] = c

    if "is_active" not in headers:
        new_col = max_col + 1
        ws.cell(row=header_row, column=new_col).value = "is_active"
        headers["is_active"] = new_col
        # tüm mevcut oyuncular için is_active=1 (varsayılan aktif)
        r = header_row + 1
        while ws.cell(row=r, column=headers["player_id"]).value:
            ws.cell(row=r, column=new_col).value = 1
            r += 1
        print("[BILGI] Players sheet'ine 'is_active' kolonu eklendi "
              "(tüm mevcut oyuncular is_active=1 olarak ayarlandı).")

    return header_row, headers


def load_rows(ws, header_row, headers):
    """Tüm oyuncu satırlarını {row_idx: {player_id, name, team, position, is_active}} olarak döndür."""
    rows = {}
    r = header_row + 1
    while ws.cell(row=r, column=headers["player_id"]).value:
        rows[r] = {
            "player_id": ws.cell(row=r, column=headers["player_id"]).value,
            "name": ws.cell(row=r, column=headers["name"]).value,
            "team": ws.cell(row=r, column=headers["team"]).value,
            "position": ws.cell(row=r, column=headers["position"]).value,
            "is_active": ws.cell(row=r, column=headers["is_active"]).value,
        }
        r += 1
    return rows


def match_player(transfer, rows, prefer_active=True):
    """İsim + takım ile en iyi eşleşmeyi bulur.
    prefer_active=True ise aktif oyunculara öncelik verilir.
    (row_idx, score) döndürür."""
    best_row, best_score = None, 0.0
    for row_idx, info in rows.items():
        score = _similarity(transfer["player_name"], info["name"])
        # takım eşleşirse küçük bonus
        if transfer.get("team") and info.get("team"):
            if transfer["team"].lower().strip() in info["team"].lower() or \
               info["team"].lower().strip() in transfer["team"].lower():
                score += 0.05
        # pasif oyuncuyu cezalandır (tercihen aktif eşleşme bul)
        if prefer_active and info.get("is_active") in (0, "0"):
            score -= 0.10
        if score > best_score:
            best_row, best_score = row_idx, score
    return best_row, best_score


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("excel_path")
    ap.add_argument("json_path")
    ap.add_argument("--apply", action="store_true",
                    help="Verilmezse sadece rapor gösterir, dosyaya yazmaz.")
    args = ap.parse_args()

    with open(args.json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(args.excel_path)
    ws = wb["Players"]
    header_row, headers = _get_or_create_headers(ws)
    rows = load_rows(ws, header_row, headers)

    # ADIM 1: veri kalitesi kontrolü
    quality = validate_transfer_window(data.get("transfers", []),
                                        total_player_count=len(rows))
    quality.print_report(f"Transfer penceresi ({data.get('transfer_date', '?')})")

    # ADIM 2: kayıtları işle — her biri için aksiyon belirle
    actions = []  # her aksiyon: {type, transfer, row, ...}
    unmatched = []

    # Yeni player_id'leri önceden hesapla — her "in" transferi için bir sonraki ID
    existing_ids = [r["player_id"] for r in rows.values()]
    next_id_counter = 0  # _next_player_id'in üzerine eklenir

    for tr in quality.valid_records:
        tt = tr["transfer_type"]
        if tt == "in":
            # yeni oyuncu — mevcut aktif oyuncuyla yüksek benzerlik varsa uyarı
            row_idx, score = match_player(tr, rows, prefer_active=True)
            if row_idx is not None and score >= NAME_MATCH_THRESHOLD:
                # zaten aktif oyuncu var — bu ya hatalı "in" ya da pasiften geri dönüş
                info = rows[row_idx]
                actions.append({
                    "type": "warn_existing",
                    "transfer": tr,
                    "matched_row": row_idx,
                    "matched_name": info["name"],
                    "matched_team": info["team"],
                    "is_active": info["is_active"],
                    "similarity": round(score, 3),
                })
            else:
                # gerçekten yeni oyuncu — yeni satır eklenecek
                # her seferinde bir sonraki ID'yi hesapla (counter artarak)
                base_n = 0
                for pid in existing_ids:
                    m = re.match(r"^PLY(\d+)$", str(pid))
                    if m:
                        n = int(m.group(1))
                        if n > base_n:
                            base_n = n
                new_id = f"PLY{base_n + 1 + next_id_counter}"
                next_id_counter += 1
                actions.append({
                    "type": "add_new",
                    "transfer": tr,
                    "new_player_id": new_id,
                })

        elif tt in ("out", "move"):
            row_idx, score = match_player(tr, rows, prefer_active=(tt == "out"))
            if row_idx is None or score < NAME_MATCH_THRESHOLD:
                unmatched.append(tr)
                continue
            info = rows[row_idx]
            if tt == "out":
                actions.append({
                    "type": "deactivate",
                    "transfer": tr,
                    "matched_row": row_idx,
                    "matched_name": info["name"],
                    "matched_team": info["team"],
                    "old_is_active": info["is_active"],
                    "similarity": round(score, 3),
                })
            else:  # move
                actions.append({
                    "type": "change_team",
                    "transfer": tr,
                    "matched_row": row_idx,
                    "matched_name": info["name"],
                    "old_team": info["team"],
                    "new_team": tr["team"],
                    "similarity": round(score, 3),
                })

    # ADIM 3: rapor bas
    print(f"\n=== Transfer penceresi ({data.get('transfer_date', '?')}) — eşleşme raporu ===")
    print(f"Geçerli kayıt: {len(quality.valid_records)}  |  "
          f"Reddedilen: {len(quality.rejected_records)}  |  "
          f"İsim eşleşmeyen: {len(unmatched)}\n")

    n_add = sum(1 for a in actions if a["type"] == "add_new")
    n_deact = sum(1 for a in actions if a["type"] == "deactivate")
    n_move = sum(1 for a in actions if a["type"] == "change_team")
    n_warn = sum(1 for a in actions if a["type"] == "warn_existing")
    print(f"  Eklenecek (in):        {n_add}")
    print(f"  Pasifleştirilecek (out): {n_deact}")
    print(f"  Takım değişikliği (move): {n_move}")
    print(f"  Uyarı (zaten var):     {n_warn}")
    print()

    if n_add > 0:
        print("--- EKLENECEK OYUNCULAR (in) ---")
        for a in actions:
            if a["type"] != "add_new":
                continue
            tr = a["transfer"]
            price = tr.get("new_price_tl", DEFAULT_NEW_PLAYER_PRICE)
            print(f"  {a['new_player_id']:8} | {tr['player_name']:<28} | "
                  f"{tr['team']:<16} | {tr['position']:<16} | "
                  f"{price:>12,.0f} TL  {tr.get('source_note', '')}".replace(",", "."))

    if n_deact > 0:
        print("\n--- PASİFLEŞTİRİLECEK OYUNCULAR (out, is_active=0) ---")
        for a in actions:
            if a["type"] != "deactivate":
                continue
            tr = a["transfer"]
            print(f"  [{a['matched_name']:<28} ({a['matched_team']})] "
                  f"is_active: {a['old_is_active']} → 0  "
                  f"(benzerlik={a['similarity']})  {tr.get('source_note', '')}")

    if n_move > 0:
        print("\n--- TAKIM DEĞİŞİKLİĞİ (move) ---")
        for a in actions:
            if a["type"] != "change_team":
                continue
            tr = a["transfer"]
            print(f"  {a['matched_name']:<28}  {a['old_team']} → {a['new_team']}  "
                  f"(benzerlik={a['similarity']})  {tr.get('source_note', '')}")

    if n_warn > 0:
        print("\n--- UYARI: 'in' olarak işaretlenen ama zaten VAR olan oyuncular ---")
        for a in actions:
            if a["type"] != "warn_existing":
                continue
            tr = a["transfer"]
            print(f"  {tr['player_name']:<28} → eşleşti: {a['matched_name']} "
                  f"({a['matched_team']}, is_active={a['is_active']})  "
                  f"(benzerlik={a['similarity']})")
            print(f"    -> EĞER bu oyuncu pasifse (is_active=0) ve geri dönüyorsa, "
                  f"transfer_type='reactivate' kullanın (henüz yok) veya manuel "
                  f"olarak is_active=1 yapın. EĞER bu yeni bir oyuncuysa, "
                  f"isim benzerliği nedeniyle karıştı — JSON'da adı belirt.")

    if unmatched:
        print("\n--- İSİM EŞLEŞMEYEN (out/move, manuel kontrol) ---")
        for u in unmatched:
            print(f"  {u['player_name']} ({u.get('team')}) - {u.get('source_note','')}")

    # ADIM 4: --apply kontrolü
    if not args.apply:
        print("\n[DRY-RUN] Hiçbir şey yazılmadı (is_active kolonu eklenmesi dahil). "
              "Onaylıyorsan --apply ile tekrar çalıştır.")
        return

    # ADIM 5: gerçek yazma
    # is_active kolonu zaten _get_or_create_headers içinde eklendi (wet-write).
    # Şimdi aksiyonları uygula:
    applied = {"add_new": 0, "deactivate": 0, "change_team": 0}

    # Son dolu satırı bul (player_id'si olan en son satır) — ws.max_row
    # bazen None dolu "hayalet" satırları sayar, gerçek son satırı bulmak lazım
    last_data_row = header_row
    r = header_row + 1
    while ws.cell(row=r, column=headers["player_id"]).value:
        last_data_row = r
        r += 1

    for a in actions:
        if a["type"] == "deactivate":
            ws.cell(row=a["matched_row"], column=headers["is_active"]).value = 0
            applied["deactivate"] += 1
        elif a["type"] == "change_team":
            ws.cell(row=a["matched_row"], column=headers["team"]).value = a["new_team"]
            applied["change_team"] += 1
        elif a["type"] == "add_new":
            # yeni satır: son dolu satırın BİR SONRASINA yaz (boşluk bırakma)
            last_data_row += 1
            new_row = last_data_row
            tr = a["transfer"]
            ws.cell(row=new_row, column=headers["player_id"]).value = a["new_player_id"]
            ws.cell(row=new_row, column=headers["name"]).value = tr["player_name"]
            ws.cell(row=new_row, column=headers["position"]).value = tr["position"]
            ws.cell(row=new_row, column=headers["team"]).value = tr["team"]
            # team_code opsiyonel — boş bırak (data_loader kullanmıyor)
            if "team_code" in headers:
                ws.cell(row=new_row, column=headers["team_code"]).value = ""
            ws.cell(row=new_row, column=headers["price_tl_gw1"]).value = tr.get(
                "new_price_tl", DEFAULT_NEW_PLAYER_PRICE)
            # play_probability default 1.0 (yeni oyuncu, sakatlık bilgisi yok)
            ws.cell(row=new_row, column=headers["play_probability"]).value = 1.0
            # last_season_* boş (uydurma veri yazma, docs/06 kuralı)
            if "is_set_piece_taker" in headers:
                ws.cell(row=new_row, column=headers["is_set_piece_taker"]).value = 0
            ws.cell(row=new_row, column=headers["is_active"]).value = 1
            applied["add_new"] += 1

    wb.save(args.excel_path)
    print(f"\n[UYGULANDI] {applied['add_new']} yeni oyuncu eklendi, "
          f"{applied['deactivate']} oyuncu pasifleştirildi, "
          f"{applied['change_team']} oyuncunun takımı güncellendi. "
          f"-> {args.excel_path}")


if __name__ == "__main__":
    main()
